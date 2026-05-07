import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkfont
import sqlite3
import hashlib
import datetime
import random
import os
import ctypes
import smtplib
import socket
from email.message import EmailMessage

try:
    import openpyxl
except ImportError:
    openpyxl = None


# 全局字体配置
BASE_FONT = ("Microsoft YaHei", 11)
TITLE_FONT = ("Microsoft YaHei", 24, "bold")
SUBTITLE_FONT = ("Microsoft YaHei", 13, "bold")


# -------------------- DPI 适配 & 数据库路径 --------------------


def set_dpi_awareness():
    """
    在 Windows 下启用高 DPI 感知，避免高分屏上界面被系统缩放变小/模糊。
    优先使用 Per-Monitor V2，其次 Per-Monitor，再其次 Process DPI Aware。
    """
    if os.name != "nt":
        return
    try:
        # Windows 10+，DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2 = -4
        awareness_context = ctypes.c_void_p(-4)
        ctypes.windll.user32.SetProcessDpiAwarenessContext(awareness_context)
    except Exception:
        try:
            # Windows 8.1 API
            ctypes.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
        except Exception:
            try:
                # 最老的方式
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass


def get_db_path():
    """
    将数据库放到 Windows 用户的 AppData\\Roaming\\PersonDrawApp 下。
    例如：C:\\Users\\用户名\\AppData\\Roaming\\PersonDrawApp\\person_draw.db
    """
    appdata = os.getenv("APPDATA")
    if not appdata:
        appdata = os.path.expanduser("~")
    app_dir = os.path.join(appdata, "PersonDrawApp")
    os.makedirs(app_dir, exist_ok=True)
    return os.path.join(app_dir, "person_draw.db")


DB_FILE = get_db_path()


def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()


# -------------------- 数据库层 --------------------


class Database:
    def __init__(self, db_path=DB_FILE):
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self.init_schema()

    def init_schema(self):
        c = self.conn.cursor()

        # 用户表
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin','user')),
                email TEXT
            )
        """
        )

        # 兼容老版本库，尝试补充 email 字段
        try:
            c.execute("ALTER TABLE users ADD COLUMN email TEXT")
        except sqlite3.OperationalError:
            pass

        # 兼容老版本库，尝试补充 receive_email 字段
        try:
            c.execute("ALTER TABLE users ADD COLUMN receive_email INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass

        # 兼容老版本库，尝试补充 password_changed 字段
        try:
            c.execute("ALTER TABLE users ADD COLUMN password_changed INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass

        # 若无用户则创建一个默认管理员
        c.execute("SELECT COUNT(*) AS cnt FROM users")
        if c.fetchone()["cnt"] == 0:
            c.execute(
                "INSERT INTO users (username,password_hash,role,email,receive_email) VALUES (?,?,?,?,?)",
                ("admin", hash_password("admin"), "admin", "", 1),
            )

        # 专家名库表
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS people (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                unit TEXT,
                phone TEXT NOT NULL,
                blocked INTEGER NOT NULL DEFAULT 0
            )
        """
        )

        # 兼容老版本库，尝试补充 unit 和 blocked 字段
        try:
            c.execute("ALTER TABLE people ADD COLUMN unit TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            c.execute("ALTER TABLE people ADD COLUMN blocked INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass

        # 抽签会话索引表（每次完整抽 3 人为一次记录）
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                created_at TEXT NOT NULL,
                completed INTEGER NOT NULL DEFAULT 0
            )
        """
        )
        
        # 为旧数据添加 completed 字段（如果不存在）
        try:
            c.execute("ALTER TABLE sessions ADD COLUMN completed INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass

        # 抽签详细日志表
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS draw_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER NOT NULL,
                person_id INTEGER NOT NULL,
                order_no INTEGER NOT NULL,
                present INTEGER NOT NULL,
                absent_reason TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(session_id) REFERENCES sessions(id),
                FOREIGN KEY(person_id) REFERENCES people(id)
            )
        """
        )
        
        # 接收邮箱列表表
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS email_recipients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT NOT NULL UNIQUE,
                name TEXT,
                created_at TEXT NOT NULL
            )
        """
        )

        self.conn.commit()
        
        # 清理未完成的会话及其日志
        self.delete_incomplete_sessions()

    # ---------- 用户相关 ----------

    def register_user(self, username, password, role="user", email="", receive_email=1):
        c = self.conn.cursor()
        try:
            c.execute(
                "INSERT INTO users (username,password_hash,role,email,receive_email) VALUES (?,?,?,?,?)",
                (username, hash_password(password), role, email, receive_email),
            )
            self.conn.commit()
            return True, "注册成功"
        except sqlite3.IntegrityError:
            return False, "用户名已存在"

    def authenticate(self, username, password):
        c = self.conn.cursor()
        c.execute("SELECT * FROM users WHERE username=?", (username,))
        row = c.fetchone()
        if not row:
            return None
        if row["password_hash"] == hash_password(password):
            return row
        return None

    def get_all_users(self):
        c = self.conn.cursor()
        c.execute("SELECT id, username, role, email, receive_email FROM users ORDER BY id ASC")
        return c.fetchall()

    def update_user(self, user_id, username, role, new_password=None, email="", receive_email=None):
        c = self.conn.cursor()
        # 检查是否存在
        c.execute("SELECT * FROM users WHERE id=?", (user_id,))
        row = c.fetchone()
        if not row:
            raise RuntimeError("用户不存在")

        # 若修改为普通用户，需保证至少还有一个管理员
        if row["role"] == "admin" and role == "user":
            c.execute("SELECT COUNT(*) AS cnt FROM users WHERE role='admin'")
            if c.fetchone()["cnt"] <= 1:
                raise RuntimeError("至少需要保留一个管理员账号")

        # 如果没有指定receive_email，保持原值
        if receive_email is None:
            receive_email = row["receive_email"] if "receive_email" in row.keys() else 1

        if new_password:
            c.execute(
                """
                UPDATE users
                SET username=?, role=?, password_hash=?, email=?, receive_email=?
                WHERE id=?
                """,
                (username, role, hash_password(new_password), email, receive_email, user_id),
            )
        else:
            c.execute(
                """
                UPDATE users
                SET username=?, role=?, email=?, receive_email=?
                WHERE id=?
                """,
                (username, role, email, receive_email, user_id),
            )
        self.conn.commit()

    def delete_user(self, user_id, current_user_id=None):
        c = self.conn.cursor()
        c.execute("SELECT * FROM users WHERE id=?", (user_id,))
        row = c.fetchone()
        if not row:
            raise RuntimeError("用户不存在")

        # 不允许删除当前登录账号，避免状态混乱
        if current_user_id is not None and user_id == current_user_id:
            raise RuntimeError("不能删除当前登录的账号")

        # 若删除管理员，需保证至少还有一个管理员
        if row["role"] == "admin":
            c.execute("SELECT COUNT(*) AS cnt FROM users WHERE role='admin'")
            if c.fetchone()["cnt"] <= 1:
                raise RuntimeError("至少需要保留一个管理员账号")

        c.execute("DELETE FROM users WHERE id=?", (user_id,))
        self.conn.commit()

    def export_users_to_excel(self, path):
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl 库，无法导出 Excel")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(["用户名", "密码(明文)", "角色(admin/user)", "电子邮箱"])
        c = self.conn.cursor()
        c.execute("SELECT username, role, email FROM users ORDER BY id ASC")
        # 密码无法还原成明文，这里导出为空，方便批量编辑再导入
        for r in c.fetchall():
            ws.append([r["username"], "", r["role"], r["email"] or ""])
        wb.save(path)

    def import_users_from_excel(self, path):
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl 库，无法导入 Excel")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or not row[0]:
                continue
            username = str(row[0]).strip()
            password = str(row[1]).strip() if len(row) > 1 and row[1] else "123456"
            role = (
                str(row[2]).strip()
                if len(row) > 2 and row[2] in ("admin", "user")
                else "user"
            )
            email = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            if not username:
                continue
            # 如果已存在则更新角色并可选重置密码，否则新建
            c = self.conn.cursor()
            c.execute("SELECT id FROM users WHERE username=?", (username,))
            exist = c.fetchone()
            if exist:
                uid = exist["id"]
                # 更新：用户名不变（按 Excel）、角色变更，如提供密码则重置
                if row[1]:
                    self.update_user(uid, username, role, new_password=password, email=email)
                else:
                    self.update_user(uid, username, role, email=email)
            else:
                self.register_user(username, password, role=role, email=email)

    # ---------- 人员表 ----------

    def get_all_people(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM people ORDER BY id ASC")
        return c.fetchall()

    def get_available_people(self):
        """获取未被屏蔽的人员列表（用于抽签）"""
        c = self.conn.cursor()
        c.execute("SELECT * FROM people WHERE blocked=0 ORDER BY id ASC")
        return c.fetchall()

    def add_person(self, name, phone, unit=""):
        c = self.conn.cursor()
        c.execute(
            "INSERT INTO people (name,unit,phone,blocked) VALUES (?,?,?,0)",
            (name, unit, phone),
        )
        self.conn.commit()

    def delete_person(self, person_id):
        c = self.conn.cursor()
        c.execute("DELETE FROM people WHERE id=?", (person_id,))
        self.conn.commit()

    def delete_all_people(self):
        c = self.conn.cursor()
        c.execute("DELETE FROM people")
        self.conn.commit()

    def delete_all_logs(self):
        """清空所有抽签日志"""
        c = self.conn.cursor()
        c.execute("DELETE FROM draw_logs")
        c.execute("DELETE FROM sessions")
        self.conn.commit()

    def reset_database(self):
        """重置所有系统数据"""
        c = self.conn.cursor()
        # 清空所有数据表
        c.execute("DELETE FROM draw_logs")
        c.execute("DELETE FROM sessions")
        c.execute("DELETE FROM people")
        c.execute("DELETE FROM users")
        # 重新创建默认管理员账户
        c.execute(
            "INSERT INTO users (username,password_hash,role,email,receive_email) VALUES (?,?,?,?,?)",
            ("admin", hash_password("admin"), "admin", "", 1),
        )
        self.conn.commit()

    def update_person(self, person_id, name, phone, unit=""):
        c = self.conn.cursor()
        c.execute(
            "UPDATE people SET name=?, unit=?, phone=? WHERE id=?",
            (name, unit, phone, person_id),
        )
        self.conn.commit()

    def set_person_blocked(self, person_id, blocked):
        """设置人员的屏蔽状态"""
        c = self.conn.cursor()
        c.execute(
            "UPDATE people SET blocked=? WHERE id=?",
            (1 if blocked else 0, person_id),
        )
        self.conn.commit()

    # ---------- 抽签 / 日志 ----------

    def create_session(self, title):
        c = self.conn.cursor()
        c.execute(
            "INSERT INTO sessions (title,created_at,completed) VALUES (?,?,0)",
            (
                title,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        self.conn.commit()
        return c.lastrowid

    def complete_session(self, session_id):
        """标记会话为已完成"""
        c = self.conn.cursor()
        c.execute(
            "UPDATE sessions SET completed=1 WHERE id=?",
            (session_id,),
        )
        self.conn.commit()

    def delete_incomplete_sessions(self):
        """删除所有未完成的会话及其日志"""
        c = self.conn.cursor()
        # 先删除未完成会话的日志
        c.execute(
            "DELETE FROM draw_logs WHERE session_id IN (SELECT id FROM sessions WHERE completed=0)"
        )
        # 再删除未完成的会话
        c.execute("DELETE FROM sessions WHERE completed=0")
        self.conn.commit()

    def add_draw_log(self, session_id, person_id, order_no, present, absent_reason):
        c = self.conn.cursor()
        c.execute(
            """
            INSERT INTO draw_logs
            (session_id,person_id,order_no,present,absent_reason,created_at)
            VALUES (?,?,?,?,?,?)
        """,
            (
                session_id,
                person_id,
                order_no,
                1 if present else 0,
                absent_reason,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        self.conn.commit()

    def get_sessions(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM sessions WHERE completed=1 ORDER BY id DESC")
        return c.fetchall()

    def get_session_logs(self, session_id):
        """获取某论证项目的所有抽签/补签日志，按记录时间倒序（最新在上）"""
        c = self.conn.cursor()
        c.execute(
            """
            SELECT dl.*, p.name, p.phone
            FROM draw_logs dl
            JOIN people p ON p.id = dl.person_id
            WHERE dl.session_id=?
            ORDER BY dl.created_at DESC, dl.id DESC
        """,
            (session_id,),
        )
        return c.fetchall()

    def get_completed_sessions(self):
        """返回已完成抽签的论证项目"""
        c = self.conn.cursor()
        c.execute("SELECT * FROM sessions WHERE completed=1 ORDER BY id DESC")
        return c.fetchall()

    def get_present_logs(self, session_id):
        """返回某论证项目中所有到场的抽签记录（用于补抽时选择标记不到场）"""
        c = self.conn.cursor()
        c.execute(
            """
            SELECT dl.*, p.name, p.phone
            FROM draw_logs dl
            JOIN people p ON p.id = dl.person_id
            WHERE dl.session_id=? AND dl.present=1
            ORDER BY dl.order_no ASC, dl.id ASC
        """,
            (session_id,),
        )
        return c.fetchall()

    def add_supplement_absent_log(self, session_id, person_id):
        """补签时新增一条“后续不到场”记录，不修改原记录"""
        order_no = self.get_max_order_no(session_id) + 1
        self.add_draw_log(
            session_id,
            person_id,
            order_no,
            present=False,
            absent_reason="专家后续不到场，进行再次抽选。",
        )

    def get_max_order_no(self, session_id):
        """获取某论证项目中当前最大的 order_no，用于补抽时续写"""
        c = self.conn.cursor()
        c.execute(
            "SELECT COALESCE(MAX(order_no), 0) AS mx FROM draw_logs WHERE session_id=?",
            (session_id,),
        )
        return c.fetchone()["mx"]

    # Excel IO - 人员
    def export_people_to_excel(self, path):
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl 库，无法导出 Excel")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "people"
        ws.append(["姓名", "单位", "电话", "是否屏蔽"])
        for row in self.get_all_people():
            ws.append([row["name"], row["unit"] or "", row["phone"], "是" if row["blocked"] else "否"])
        wb.save(path)

    def import_people_from_excel(self, path):
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl 库，无法导入 Excel")
        wb = openpyxl.load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                values = list(row) if row else []
                if len(values) < 9:
                    continue
                name = values[1]
                unit = values[7]
                phone = values[8]
                if not name or not phone:
                    continue
                name = str(name).strip()
                unit = str(unit).strip() if unit else ""
                phone = str(phone).strip()
                if name in ("姓名", "") or phone in ("联系电话", ""):
                    continue
                c = self.conn.cursor()
                c.execute("SELECT id FROM people WHERE name=? AND phone=?", (name, phone))
                exist = c.fetchone()
                if exist:
                    c.execute(
                        "UPDATE people SET unit=? WHERE id=?",
                        (unit, exist["id"]),
                    )
                    self.conn.commit()
                else:
                    c.execute(
                        "INSERT INTO people (name,unit,phone,blocked) VALUES (?,?,?,0)",
                        (name, unit, phone),
                    )
                    self.conn.commit()

    def get_mail_config(self):
        """获取邮件配置（若表不存在则初始化）"""
        c = self.conn.cursor()
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS mail_config (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                smtp_server TEXT,
                smtp_port INTEGER,
                use_ssl INTEGER,
                use_tls INTEGER,
                username TEXT,
                password TEXT,
                from_addr TEXT
            )
        """
        )
        c.execute("INSERT OR IGNORE INTO mail_config (id) VALUES (1)")
        c.execute("SELECT * FROM mail_config WHERE id=1")
        row = c.fetchone()
        return {
            "smtp_server": row["smtp_server"],
            "smtp_port": row["smtp_port"] or 0,
            "use_ssl": bool(row["use_ssl"]),
            "use_tls": bool(row["use_tls"]),
            "username": row["username"],
            "password": row["password"],
            "from_addr": row["from_addr"],
        }

    def save_mail_config(self, cfg: dict):
        c = self.conn.cursor()
        c.execute(
            """
            UPDATE mail_config
            SET smtp_server=?, smtp_port=?, use_ssl=?, use_tls=?, username=?, password=?, from_addr=?
            WHERE id=1
        """,
            (
                cfg.get("smtp_server") or "",
                int(cfg.get("smtp_port") or 0),
                1 if cfg.get("use_ssl") else 0,
                1 if cfg.get("use_tls") else 0,
                cfg.get("username") or "",
                cfg.get("password") or "",
                cfg.get("from_addr") or "",
            ),
        )
        self.conn.commit()

    # ---------- 邮箱列表管理 ----------
    def get_email_recipients(self):
        """获取所有接收邮箱列表"""
        c = self.conn.cursor()
        c.execute("SELECT * FROM email_recipients ORDER BY id ASC")
        return c.fetchall()

    def add_email_recipient(self, email, name=""):
        """添加接收邮箱"""
        c = self.conn.cursor()
        c.execute(
            "INSERT INTO email_recipients (email, name, created_at) VALUES (?,?,?)",
            (email, name, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )
        self.conn.commit()

    def update_email_recipient(self, recipient_id, email, name=""):
        """更新接收邮箱"""
        c = self.conn.cursor()
        c.execute(
            "UPDATE email_recipients SET email=?, name=? WHERE id=?",
            (email, name, recipient_id),
        )
        self.conn.commit()

    def delete_email_recipient(self, recipient_id):
        """删除接收邮箱"""
        c = self.conn.cursor()
        c.execute("DELETE FROM email_recipients WHERE id=?", (recipient_id,))
        self.conn.commit()

    def get_recipient_emails(self):
        """获取所有接收邮箱地址列表"""
        c = self.conn.cursor()
        c.execute("SELECT email FROM email_recipients")
        return [r["email"] for r in c.fetchall()]

    # ---------- Excel IO - 日志 ----------
    def export_logs_to_excel(self, path):
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl 库，无法导出 Excel")
        wb = openpyxl.Workbook()

        # Sheet 1: sessions
        ws1 = wb.active
        ws1.title = "sessions"
        ws1.append(["ID", "标题", "创建时间"])
        for s in self.get_sessions():
            ws1.append([s["id"], s["title"], s["created_at"]])

        # Sheet 2: draw_logs
        ws2 = wb.create_sheet("draw_logs")
        ws2.append(
            [
                "ID",
                "SessionID",
                "抽签顺序",
                "人员ID",
                "姓名",
                "手机号",
                "到场情况",
                "备注",
                "记录时间",
            ]
        )
        c = self.conn.cursor()
        c.execute(
            """
            SELECT dl.*, p.name, p.phone
            FROM draw_logs dl
            JOIN people p ON p.id = dl.person_id
            ORDER BY dl.session_id DESC, dl.created_at DESC, dl.id DESC
        """
        )

        def present_display(row):
            if row["present"] == 1:
                return "到场"
            if row["absent_reason"] == "专家后续不到场，进行再次抽选。":
                return "后续不到场"
            return "不到场"

        for r in c.fetchall():
            ws2.append(
                [
                    r["id"],
                    r["session_id"],
                    r["order_no"],
                    r["person_id"],
                    r["name"],
                    r["phone"],
                    present_display(r),
                    r["absent_reason"] or "",
                    r["created_at"],
                ]
            )

        wb.save(path)


# -------------------- 各个界面 Frame --------------------


class LoginFrame(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.build_ui()

    def build_ui(self):
        container = ttk.Frame(self, padding=80)
        container.place(relx=0.5, rely=0.5, anchor="center")

        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=1)

        # Logo图片
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "珠海安防协会logo.gif")
            if os.path.exists(logo_path):
                logo_img = tk.PhotoImage(file=logo_path)
                logo_img = logo_img.subsample(7, 7)  # 缩小到1/7
                logo_label = ttk.Label(container, image=logo_img)
                logo_label.image = logo_img
                logo_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        except Exception:
            pass

        title = ttk.Label(
            container,
            text="珠海安防协会项目论证\n专家抽签系统",
            font=TITLE_FONT,
            justify="center",
        )
        title.grid(row=1, column=0, columnspan=2, pady=(0, 50))

        ttk.Label(container, text="用户名:").grid(
            row=2, column=0, sticky="e", padx=20, pady=12
        )
        self.username_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.username_var, width=28).grid(
            row=2, column=1, sticky="w", padx=20, pady=12
        )

        ttk.Label(container, text="密码:").grid(
            row=3, column=0, sticky="e", padx=20, pady=12
        )
        self.password_var = tk.StringVar()
        self.pwd_entry = ttk.Entry(container, textvariable=self.password_var, show="*", width=28)
        self.pwd_entry.grid(
            row=3, column=1, sticky="w", padx=20, pady=12
        )
        
        # 显示密码复选框
        self.show_pwd_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            container, text="显示密码", variable=self.show_pwd_var, 
            command=self._toggle_password_visibility
        ).grid(row=4, column=1, sticky="w", padx=20, pady=5)


        btn_frame = ttk.Frame(container)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="登录", command=self.login, width=14).grid(
            row=0, column=0, padx=15
        )

        ttk.Label(
            container,
            text="Copyright © 2026 Takasak1Yu",
            foreground="gray",
        ).grid(row=6, column=0, columnspan=2, pady=(20, 0))

    def _toggle_password_visibility(self):
        """切换密码显示/隐藏"""
        if self.show_pwd_var.get():
            self.pwd_entry.config(show="")
        else:
            self.pwd_entry.config(show="*")

    def login(self):
        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        if not username or not password:
            messagebox.showwarning("提示", "请输入用户名和密码")
            return
        user = self.app.db.authenticate(username, password)
        if user:
            self.app.current_user = user
            self.app.show_main_frame()
        else:
            messagebox.showerror("错误", "用户名或密码错误")


class RegisterFrame(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.build_ui()

    def build_ui(self):
        container = ttk.Frame(self, padding=80)
        container.place(relx=0.5, rely=0.5, anchor="center")

        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=1)

        ttk.Label(
            container, text="注册", font=TITLE_FONT
        ).grid(row=0, column=0, columnspan=2, pady=(0, 50))

        ttk.Label(container, text="用户名:").grid(
            row=1, column=0, sticky="e", padx=20, pady=12
        )
        self.username_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.username_var, width=28).grid(
            row=1, column=1, sticky="w", padx=20, pady=12
        )

        ttk.Label(container, text="密码:").grid(
            row=2, column=0, sticky="e", padx=20, pady=12
        )
        self.password_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.password_var, show="*", width=28).grid(
            row=2, column=1, sticky="w", padx=20, pady=12
        )

        ttk.Label(container, text="确认密码:").grid(
            row=3, column=0, sticky="e", padx=20, pady=12
        )
        self.password2_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.password2_var, show="*", width=28).grid(
            row=3, column=1, sticky="w", padx=20, pady=12
        )

        btn_frame = ttk.Frame(container)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=30)

        ttk.Button(btn_frame, text="注册", command=self.register, width=14).grid(
            row=0, column=0, padx=15
        )
        ttk.Button(btn_frame, text="返回登录", command=self.back_login, width=14).grid(
            row=0, column=1, padx=15
        )

    def register(self):
        u = self.username_var.get().strip()
        p1 = self.password_var.get().strip()
        p2 = self.password2_var.get().strip()
        if not u or not p1 or not p2:
            messagebox.showwarning("提示", "请完整填写信息")
            return
        if p1 != p2:
            messagebox.showwarning("提示", "两次密码不一致")
            return
        ok, msg = self.app.db.register_user(u, p1, role="user")
        if ok:
            messagebox.showinfo("成功", msg)
            self.back_login()
        else:
            messagebox.showerror("失败", msg)

    def back_login(self):
        self.app.show_login_frame()


class PeopleFrame(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.build_ui()
        self.refresh()

    def build_ui(self):
        outer = ttk.Frame(self, padding=30)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x", pady=(0, 15))

        ttk.Label(top, text="专家名库", font=SUBTITLE_FONT).pack(
            side="left", padx=8
        )

        # 标题下方中部按钮栏
        btn_bar = ttk.Frame(outer)
        btn_bar.pack(pady=15)

        self.btn_add = ttk.Button(btn_bar, text="新增", command=self.add_person, width=9)
        self.btn_edit = ttk.Button(btn_bar, text="编辑", command=self.edit_person, width=9)
        self.btn_del = ttk.Button(btn_bar, text="删除", command=self.delete_person, width=9)
        self.btn_block = ttk.Button(btn_bar, text="屏蔽", command=self.block_person, width=9)
        self.btn_unblock = ttk.Button(btn_bar, text="取消屏蔽", command=self.unblock_person, width=11)
        self.btn_import = ttk.Button(
            btn_bar, text="从Excel导入", command=self.import_excel, width=11
        )
        self.btn_export = ttk.Button(
            btn_bar, text="导出至Excel", command=self.export_excel, width=11
        )

        for b in (
                self.btn_add,
                self.btn_edit,
                self.btn_del,
                self.btn_block,
                self.btn_unblock,
                self.btn_import,
                self.btn_export,
        ):
            b.pack(side="left", padx=8)






        table_frame = ttk.Frame(outer)
        table_frame.pack(fill="both", expand=True)

        columns = ("id", "name", "unit", "phone", "blocked")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=18,
        )
        self.tree.heading("id", text="")
        self.tree.heading("name", text="姓名")
        self.tree.heading("unit", text="单位")
        self.tree.heading("phone", text="电话")
        self.tree.heading("blocked", text="是否屏蔽")
        self.tree.column("id", width=0, minwidth=0, anchor="center", stretch=False)
        self.tree.column("name", width=120, anchor="center", stretch=True)
        self.tree.column("unit", width=280, anchor="center", stretch=True)
        self.tree.column("phone", width=140, anchor="center", stretch=True)
        self.tree.column("blocked", width=80, anchor="center", stretch=False)
        self.tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")



    def set_admin_mode(self, is_admin):
        state = "normal" if is_admin else "disabled"
        for b in (self.btn_add, self.btn_edit, self.btn_del, self.btn_block, self.btn_unblock, self.btn_import):
            b["state"] = state
        self.btn_export["state"] = "normal"

    def refresh(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for row in self.app.db.get_all_people():
            blocked_text = "是" if row["blocked"] else "否"
            self.tree.insert(
                "",
                "end",
                values=(row["id"], row["name"], row["unit"] or "", row["phone"], blocked_text),
            )
        self.app.auto_adjust_columns(self.tree)

    def _get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        return int(self.tree.item(sel[0], "values")[0])

    def add_person(self):
        self._edit_dialog()

    def edit_person(self):
        pid = self._get_selected_id()
        if not pid:
            messagebox.showwarning("提示", "请先选择一条记录")
            return
        people = {p["id"]: p for p in self.app.db.get_all_people()}
        if pid not in people:
            return
        p = people[pid]
        self._edit_dialog(pid, p["name"], p["phone"], p["unit"] or "")

    def block_person(self):
        pid = self._get_selected_id()
        if not pid:
            messagebox.showwarning("提示", "请先选择一条记录")
            return
        self.app.db.set_person_blocked(pid, True)
        self.refresh()
        messagebox.showinfo("成功", "已屏蔽该人员，将不参与抽签")

    def unblock_person(self):
        pid = self._get_selected_id()
        if not pid:
            messagebox.showwarning("提示", "请先选择一条记录")
            return
        self.app.db.set_person_blocked(pid, False)
        self.refresh()
        messagebox.showinfo("成功", "已取消屏蔽该人员")

    def _edit_dialog(self, person_id=None, name="", phone="", unit=""):
        dlg = tk.Toplevel(self)
        dlg.title("人员信息")
        dlg.grab_set()
        dlg.resizable(False, False)

        container = ttk.Frame(dlg, padding=25)
        container.grid(row=0, column=0, sticky="nsew")
        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)

        ttk.Label(container, text="姓名:").grid(
            row=0, column=0, padx=8, pady=10, sticky="e"
        )
        name_var = tk.StringVar(value=name)
        ttk.Entry(container, textvariable=name_var, width=28).grid(
            row=0, column=1, padx=8, pady=10, sticky="w"
        )

        ttk.Label(container, text="单位:").grid(
            row=1, column=0, padx=8, pady=10, sticky="e"
        )
        unit_var = tk.StringVar(value=unit)
        ttk.Entry(container, textvariable=unit_var, width=28).grid(
            row=1, column=1, padx=8, pady=10, sticky="w"
        )

        ttk.Label(container, text="电话:").grid(
            row=2, column=0, padx=8, pady=10, sticky="e"
        )
        phone_var = tk.StringVar(value=phone)
        ttk.Entry(container, textvariable=phone_var, width=28).grid(
            row=2, column=1, padx=8, pady=10, sticky="w"
        )

        def on_ok():
            n = name_var.get().strip()
            u = unit_var.get().strip()
            ph = phone_var.get().strip()
            if not n:
                messagebox.showwarning("提示", "姓名不能为空")
                return
            if person_id is None:
                self.app.db.add_person(n, ph, u)
            else:
                self.app.db.update_person(person_id, n, ph, u)
            self.refresh()
            dlg.destroy()

        btn_frame = ttk.Frame(container)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=(18, 0))
        ttk.Button(btn_frame, text="确定", command=on_ok, width=11).grid(
            row=0, column=0, padx=10
        )
        ttk.Button(btn_frame, text="取消", command=dlg.destroy, width=11).grid(
            row=0, column=1, padx=10
        )

    def delete_person(self):
        pid = self._get_selected_id()
        if not pid:
            messagebox.showwarning("提示", "请先选择一条记录")
            return
        if messagebox.askyesno("确认", "确定要删除该人员吗？"):
            self.app.db.delete_person(pid)
            self.refresh()

    def import_excel(self):
        if openpyxl is None:
            messagebox.showerror("错误", "请先安装 openpyxl 库")
            return
        
        # 创建自定义对话框
        dialog = tk.Toplevel(self)
        dialog.title("导入选项")
        dialog.geometry("380x280")
        dialog.resizable(False, False)
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # 添加标签
        label = ttk.Label(dialog, text="请选择导入方式：", font=("SimHei", 11))
        label.pack(pady=15)
        
        # 存储用户选择
        self.import_option = None
        
        # 处理右上角关闭按钮
        def on_close():
            self.import_option = "cancel"
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        # 创建按钮框架
        button_frame = ttk.Frame(dialog)
        button_frame.pack(expand=True, fill="both", padx=30)
        
        # 删除当前名库重新导入按钮
        def delete_and_import():
            self.import_option = "delete"
            dialog.destroy()
        
        btn_delete = ttk.Button(
            button_frame, 
            text="删除当前名库重新导入", 
            command=delete_and_import,
            width=22
        )
        btn_delete.pack(pady=10, fill="x")
        
        # 保留当前名库继续导入按钮
        def keep_and_import():
            self.import_option = "keep"
            dialog.destroy()
        
        btn_keep = ttk.Button(
            button_frame, 
            text="保留当前名库继续导入", 
            command=keep_and_import,
            width=22
        )
        btn_keep.pack(pady=10, fill="x")
        
        # 取消按钮
        def cancel_import():
            self.import_option = "cancel"
            dialog.destroy()
        
        btn_cancel = ttk.Button(
            button_frame, 
            text="取消", 
            command=cancel_import,
            width=22
        )
        btn_cancel.pack(pady=10, fill="x")
        
        # 等待用户选择
        dialog.transient(self)
        dialog.grab_set()
        self.wait_window(dialog)
        
        # 根据用户选择执行操作
        if self.import_option != "delete" and self.import_option != "keep":
            return
        
        # 选择Excel文件
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        if not path:
            return
        
        try:
            if self.import_option == "delete":
                # 删除当前名库
                self.app.db.delete_all_people()
            
            self.app.db.import_people_from_excel(path)
            self.refresh()
            messagebox.showinfo("成功", "导入完成")
        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {e}")

    def export_excel(self):
        if openpyxl is None:
            messagebox.showerror("错误", "请先安装 openpyxl 库")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            title="导出专家名库",
            initialfile="people.xlsx",
        )
        if not path:
            return
        try:
            self.app.db.export_people_to_excel(path)
            messagebox.showinfo("成功", "导出完成")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")


class LogsFrame(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.build_ui()
        self.refresh_sessions()

    def build_ui(self):
        outer = ttk.Frame(self, padding=30)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x", pady=(0, 15))

        ttk.Label(top, text="抽签日志", font=SUBTITLE_FONT).pack(
            side="left", padx=8
        )

        # 标题下方中部按钮栏
        btn_bar = ttk.Frame(outer)
        btn_bar.pack(pady=(0, 10))
        ttk.Button(btn_bar, text="刷新", command=self.refresh_sessions, width=9).pack(
            side="left", padx=5
        )
        ttk.Button(
            btn_bar, text="发送记录至邮箱", command=self.send_logs_email, width=14
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_bar, text="导出所有日志", command=self.export_all_logs, width=12
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_bar, text="导出选择日志", command=self.export_selected_logs, width=12
        ).pack(side="left", padx=5)

        main = ttk.Panedwindow(outer, orient="horizontal")
        main.pack(fill="both", expand=True)

        # 左侧：session 列表
        left = ttk.Frame(main, padding=(0, 0, 8, 0))
        columns = ("id", "created_at", "title")
        self.sessions_tree = ttk.Treeview(
            left, columns=columns, show="headings", height=18
        )
        # 隐藏 ID 列，仅用于内部
        self.sessions_tree.heading("id", text="")
        self.sessions_tree.heading("created_at", text="时间")
        self.sessions_tree.heading("title", text="论证项目名称")
        self.sessions_tree.column("id", width=0, minwidth=0, anchor="center", stretch=False)
        self.sessions_tree.column("created_at", width=200, anchor="center", stretch=False)
        self.sessions_tree.column("title", width=260, stretch=True)
        self.sessions_tree.pack(side="left", fill="both", expand=True)

        vsb1 = ttk.Scrollbar(left, orient="vertical", command=self.sessions_tree.yview)
        self.sessions_tree.configure(yscrollcommand=vsb1.set)
        vsb1.pack(side="right", fill="y")

        self.sessions_tree.bind("<<TreeviewSelect>>", self.on_session_select)

        # 右侧：详细日志（隐藏顺序，第一列为记录时间，按最新在上排序）
        right = ttk.Frame(main, padding=(8, 0, 0, 0))
        columns2 = (
            "created_at",
            "name",
            "phone",
            "present",
            "reason",
        )
        self.logs_tree = ttk.Treeview(
            right, columns=columns2, show="headings", height=18
        )
        self.logs_tree.heading("created_at", text="记录时间")
        self.logs_tree.heading("name", text="姓名")
        self.logs_tree.heading("phone", text="手机号")
        self.logs_tree.heading("present", text="到场情况")
        self.logs_tree.heading("reason", text="备注")

        self.logs_tree.column("created_at", width=180, anchor="center", stretch=False)
        self.logs_tree.column("name", width=120, anchor="center", stretch=False)
        self.logs_tree.column("phone", width=160, anchor="center", stretch=False)
        self.logs_tree.column("present", width=100, anchor="center", stretch=False)
        self.logs_tree.column("reason", width=340, stretch=True)
        self.logs_tree.pack(side="left", fill="both", expand=True)

        vsb2 = ttk.Scrollbar(right, orient="vertical", command=self.logs_tree.yview)
        self.logs_tree.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side="right", fill="y")

        main.add(left, weight=1)
        main.add(right, weight=1)

    def refresh_sessions(self):
        for i in self.sessions_tree.get_children():
            self.sessions_tree.delete(i)
        for s in self.app.db.get_sessions():
            self.sessions_tree.insert(
                "",
                "end",
                values=(s["id"], s["created_at"], s["title"]),
            )
        for i in self.logs_tree.get_children():
            self.logs_tree.delete(i)
        self.app.auto_adjust_columns(self.sessions_tree)

    def _get_present_display(self, r):
        """到场情况显示：到场 / 后续不到场 / 不到场"""
        if r["present"] == 1:
            return "到场"
        if r["absent_reason"] == "专家后续不到场，进行再次抽选。":
            return "后续不到场"
        return "不到场"

    def on_session_select(self, event=None):
        sel = self.sessions_tree.selection()
        if not sel:
            return
        sid = int(self.sessions_tree.item(sel[0], "values")[0])
        logs = self.app.db.get_session_logs(sid)
        for i in self.logs_tree.get_children():
            self.logs_tree.delete(i)
        for r in logs:
            present_text = self._get_present_display(r)
            self.logs_tree.insert(
                "",
                "end",
                values=(
                    r["created_at"],
                    r["name"],
                    r["phone"],
                    present_text,
                    r["absent_reason"] or "",
                ),
            )
        self.app.auto_adjust_columns(self.logs_tree)

    def export_all_logs(self):
        """导出所有日志为txt文件"""
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt")],
            title="导出所有日志",
            initialfile="all_logs.txt",
        )
        if not path:
            return
        try:
            lines = []
            sessions = self.app.db.get_sessions()
            for s in sessions:
                lines.append("=" * 60)
                lines.append(f"论证项目名称: {s['title'] or ''}")
                lines.append(f"创建时间: {s['created_at']}")
                lines.append("-" * 40)
                logs = self.app.db.get_session_logs(s["id"])
                for r in logs:
                    if r["present"] == 1:
                        state = "到场"
                    elif r["absent_reason"] == "专家后续不到场，进行再次抽选。":
                        state = "后续不到场"
                    else:
                        state = "不到场"
                    lines.append(f"[{r['created_at']}] 姓名:{r['name']} 手机:{r['phone']} 到场情况:{state} 备注:{r['absent_reason'] or ''}")
                lines.append("")
            lines.append("=" * 60)
            
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            messagebox.showinfo("成功", "导出完成")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def export_selected_logs(self):
        """导出选择的日志为txt文件"""
        sel = self.sessions_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个论证项目")
            return
        if len(sel) > 1:
            messagebox.showwarning("提示", "一次只能导出一个论证项目的日志")
            return
        
        session_id = int(self.sessions_tree.item(sel[0], "values")[0])
        sessions = [s for s in self.app.db.get_sessions() if s["id"] == session_id]
        if not sessions:
            messagebox.showerror("错误", "未找到该论证项目")
            return
        s = sessions[0]
        
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt")],
            title="导出选择日志",
            initialfile=f"{s['title'] or 'logs'}.txt",
        )
        if not path:
            return
        try:
            lines = []
            lines.append("=" * 60)
            lines.append(f"论证项目名称: {s['title'] or ''}")
            lines.append(f"创建时间: {s['created_at']}")
            lines.append("-" * 40)
            logs = self.app.db.get_session_logs(s["id"])
            for r in logs:
                if r["present"] == 1:
                    state = "到场"
                elif r["absent_reason"] == "专家后续不到场，进行再次抽选。":
                    state = "后续不到场"
                else:
                    state = "不到场"
                lines.append(f"[{r['created_at']}] 姓名:{r['name']} 手机:{r['phone']} 到场情况:{state} 备注:{r['absent_reason'] or ''}")
            lines.append("")
            lines.append("=" * 60)
            
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            messagebox.showinfo("成功", "导出完成")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def send_logs_email(self):
        """从日志界面发送一个或多个论证项目的日志到所有管理员邮箱"""
        sel = self.sessions_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择至少一个论证项目")
            return
        session_ids = [int(self.sessions_tree.item(i, "values")[0]) for i in sel]
        self.app.send_sessions_email(self, session_ids)


class DrawFrame(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.current_session_id = None
        self.present_count = 0
        self.current_person = None
        self.order_no = 0
        self.people_cache = []
        self.drawn_person_ids = set()
        self.view_mode = "choice"
        self.supplement_session_id = None
        self.supp_new_session_id = None
        self.supplement_vacant_count = 0
        self.supplement_order_base = 0
        self.is_drawing = False
        self.draw_animation_id = None
        self.supp_is_drawing = False
        self.supp_animation_id = None
        self.build_ui()

    def build_ui(self):
        self.outer = ttk.Frame(self, padding=30)
        self.outer.pack(fill="both", expand=True)

        # ---------- 功能选择界面（居中垂直两选项） ----------
        self.choice_frame = ttk.Frame(self.outer)
        choice_inner = ttk.Frame(self.choice_frame, padding=60)
        choice_inner.place(relx=0.5, rely=0.5, anchor="center")

        ttk.Label(
            choice_inner,
            text="请选择抽签功能",
            font=SUBTITLE_FONT,
        ).pack(pady=(0, 40))

        ttk.Button(
            choice_inner,
            text="新论证项目抽签",
            command=self._show_new_draw,
            width=24,
        ).pack(pady=20)

        ttk.Button(
            choice_inner,
            text="过往论证项目补抽",
            command=self._show_supplement_select,
            width=24,
        ).pack(pady=20)

        # ---------- 新论证项目抽签界面 ----------
        self.new_draw_frame = ttk.Frame(self.outer)
        self._build_new_draw_ui()

        # ---------- 过往论证项目补抽界面 ----------
        self.supplement_frame = ttk.Frame(self.outer)
        self._build_supplement_ui()

        self._show_view("choice")

    def _show_view(self, mode):
        self.view_mode = mode
        for f in (self.choice_frame, self.new_draw_frame, self.supplement_frame):
            f.pack_forget()
        if mode == "choice":
            self.choice_frame.pack(fill="both", expand=True)
        elif mode == "new":
            self.new_draw_frame.pack(fill="both", expand=True)
        else:
            self.supplement_frame.pack(fill="both", expand=True)

    def _show_choice(self):
        self._show_view("choice")

    def _show_new_draw(self):
        # 弹出对话框让用户输入项目名称和申请论证单位
        dialog = tk.Toplevel(self)
        dialog.title("新建论证项目")
        dialog.geometry("450x280")
        dialog.resizable(False, False)
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # 存储用户输入
        self.new_project_name = None
        self.new_project_unit = None
        
        # 处理右上角关闭按钮
        def on_close():
            self.new_project_name = None
            self.new_project_unit = None
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        # 论证项目名称
        ttk.Label(dialog, text="请输入论证项目名称：", font=("SimHei", 11)).pack(pady=(15, 5))
        name_var = tk.StringVar()
        entry_name = ttk.Entry(dialog, textvariable=name_var, width=40)
        entry_name.pack(pady=5, padx=30)
        entry_name.focus_set()
        
        # 申请论证单位
        ttk.Label(dialog, text="请输入申请论证单位：", font=("SimHei", 11)).pack(pady=(10, 5))
        unit_var = tk.StringVar()
        entry_unit = ttk.Entry(dialog, textvariable=unit_var, width=40)
        entry_unit.pack(pady=5, padx=30)
        
        # 按钮框架
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        def on_confirm():
            self.new_project_name = name_var.get().strip() or "未命名论证项目"
            self.new_project_unit = unit_var.get().strip() or "未指定单位"
            dialog.destroy()
        
        def on_cancel():
            self.new_project_name = None
            self.new_project_unit = None
            dialog.destroy()
        
        ttk.Button(btn_frame, text="确定", command=on_confirm, width=12).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="取消", command=on_cancel, width=12).pack(side="left", padx=10)
        
        # 回车键确认
        entry_name.bind("<Return>", lambda e: entry_unit.focus_set())
        entry_unit.bind("<Return>", lambda e: on_confirm())
        
        # 等待用户操作
        dialog.transient(self)
        dialog.grab_set()
        self.wait_window(dialog)
        
        # 用户取消
        if self.new_project_name is None:
            return
        
        # 显示抽签界面
        self._show_view("new")
        self.title_var.set(self.new_project_name)
        self.unit_var.set(self.new_project_unit)
        self.present_var.set("0 / 3")
        self.name_var.set("")
        self.phone_var.set("")
        self.unit_display_var.set("")
        self.set_buttons_state(started=False)
        self.current_session_id = None
        
        # 自动开始新抽签
        self._auto_start_session()

    def _show_supplement_select(self):
        self._show_view("supplement")
        self._supplement_step = 1
        self.supplement_session_id = None
        self._refresh_supplement_session_list()
        self._show_supplement_step1()

    def _build_new_draw_ui(self):
        outer = self.new_draw_frame
        
        # 信息区域（居中显示）
        info_frame = ttk.Frame(outer)
        info_frame.pack(pady=30)
        
        # 论证项目名称（同一行显示）
        self.title_var = tk.StringVar()
        title_row = ttk.Frame(info_frame)
        title_row.pack(pady=5)
        ttk.Label(title_row, text="论证项目名称：", font=("SimHei", 12)).pack(side="left")
        ttk.Label(title_row, textvariable=self.title_var, font=("SimHei", 12, "bold")).pack(side="left", padx=5)
        
        # 申请论证单位（同一行显示）
        self.unit_var = tk.StringVar()
        unit_row = ttk.Frame(info_frame)
        unit_row.pack(pady=5)
        ttk.Label(unit_row, text="申请论证单位：", font=("SimHei", 12)).pack(side="left")
        ttk.Label(unit_row, textvariable=self.unit_var, font=("SimHei", 12, "bold")).pack(side="left", padx=5)
        
        # 已到场专家（同一行显示）
        self.present_var = tk.StringVar(value="0 / 3")
        present_row = ttk.Frame(info_frame)
        present_row.pack(pady=5)
        ttk.Label(present_row, text="已到场专家：", font=("SimHei", 12)).pack(side="left")
        ttk.Label(present_row, textvariable=self.present_var, font=("SimHei", 12, "bold")).pack(side="left", padx=5)

        ttk.Separator(outer, orient="horizontal").pack(fill="x", pady=8)

        # 抽中人员信息
        person_frame = ttk.Frame(outer)
        person_frame.pack(pady=20)

        person_frame.columnconfigure(0, weight=1)
        person_frame.columnconfigure(1, weight=1)

        ttk.Label(person_frame, text="专家姓名:").grid(
            row=0, column=0, sticky="e", padx=20, pady=10
        )
        self.name_var = tk.StringVar()
        ttk.Label(person_frame, textvariable=self.name_var, font=("Microsoft YaHei", 15)).grid(
            row=0, column=1, sticky="w", padx=20, pady=10
        )

        ttk.Label(person_frame, text="手机号:").grid(
            row=1, column=0, sticky="e", padx=20, pady=10
        )
        self.phone_var = tk.StringVar()
        ttk.Label(person_frame, textvariable=self.phone_var).grid(
            row=1, column=1, sticky="w", padx=20, pady=10
        )
        
        ttk.Label(person_frame, text="单位:").grid(
            row=2, column=0, sticky="e", padx=20, pady=10
        )
        self.unit_display_var = tk.StringVar()
        ttk.Label(person_frame, textvariable=self.unit_display_var).grid(
            row=2, column=1, sticky="w", padx=20, pady=10
        )

        # 按钮区域
        btns = ttk.Frame(outer)
        btns.pack(pady=20)

        self.btn_draw = ttk.Button(btns, text="抽签", command=self.start_draw_animation, width=12)
        self.btn_stop = ttk.Button(btns, text="停止", command=self.stop_draw_animation, width=12)
        self.btn_present = ttk.Button(
            btns, text="到场", command=lambda: self.mark_result(True), width=12
        )
        self.btn_absent = ttk.Button(
            btns, text="不到场", command=lambda: self.mark_result(False), width=12
        )
        self.btn_cancel = ttk.Button(btns, text="取消", command=self._cancel_draw, width=12)

        self.btn_draw.grid(row=0, column=0, padx=10)
        self.btn_stop.grid(row=0, column=0, padx=10)
        self.btn_present.grid(row=0, column=0, padx=10)
        self.btn_absent.grid(row=0, column=1, padx=10)
        self.btn_cancel.grid(row=0, column=2, padx=10)

        ttk.Label(
            outer,
            text="说明：一次抽签流程中，将依次确认 3 名到场人员。不到场需填写理由并可继续抽下一位。",
            foreground="gray",
        ).pack(pady=20)

        self.set_buttons_state(started=False)

    def _build_supplement_ui(self):
        outer = self.supplement_frame

        self.supp_step1 = ttk.Frame(outer)
        ttk.Label(self.supp_step1, text="选择过往论证项目", font=SUBTITLE_FONT).pack(
            anchor="w", padx=5, pady=(0, 10)
        )
        supp_table_frame = ttk.Frame(self.supp_step1)
        supp_table_frame.pack(fill="both", expand=True)

        cols = ("id", "title", "created_at")
        self.supp_sessions_tree = ttk.Treeview(
            supp_table_frame, columns=cols, show="headings", height=12
        )
        self.supp_sessions_tree.heading("id", text="ID")
        self.supp_sessions_tree.heading("title", text="论证项目名称")
        self.supp_sessions_tree.heading("created_at", text="创建时间")
        self.supp_sessions_tree.column("id", width=70)
        self.supp_sessions_tree.column("title", width=300)
        self.supp_sessions_tree.column("created_at", width=200)
        self.supp_sessions_tree.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(supp_table_frame, orient="vertical", command=self.supp_sessions_tree.yview)
        self.supp_sessions_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        supp_btn1 = ttk.Frame(self.supp_step1)
        supp_btn1.pack(pady=10)
        ttk.Button(supp_btn1, text="选择并继续", command=self._supplement_on_session_selected, width=14).pack(side="right", padx=5)
        ttk.Button(supp_btn1, text="返回", command=self._show_choice, width=10).pack(side="right")

        self.supp_step2 = ttk.Frame(outer)
        ttk.Label(self.supp_step2, text="请选择未能到场的专家", font=SUBTITLE_FONT).pack(
            anchor="center", padx=5, pady=(0, 10)
        )
        self.supp_check_vars = {}
        self.supp_check_frame = ttk.Frame(self.supp_step2)
        self.supp_check_frame.pack(fill="both", expand=True, padx=10, pady=10)

        supp_btn2 = ttk.Frame(self.supp_step2)
        supp_btn2.pack(pady=10)
        ttk.Button(supp_btn2, text="确认标记（自动新增后续不到场记录）", command=self._supplement_confirm_absent, width=26).pack(side="right", padx=5)
        ttk.Button(supp_btn2, text="上一步", command=self._show_supplement_step1, width=10).pack(side="right")

        self.supp_step3 = ttk.Frame(outer)
        ttk.Label(self.supp_step3, text="补录专家", font=SUBTITLE_FONT).pack(
            anchor="center", padx=5, pady=(0, 10)
        )
        supp_info = ttk.Frame(self.supp_step3)
        supp_info.pack(pady=10)
        supp_info.columnconfigure(0, weight=1)
        supp_info.columnconfigure(1, weight=1)
        ttk.Label(supp_info, text="当前状态:").grid(row=0, column=0, sticky="e", padx=10, pady=5)
        self.supp_status_var = tk.StringVar()
        ttk.Label(supp_info, textvariable=self.supp_status_var).grid(row=0, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(supp_info, text="已补录专家数:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
        self.supp_present_var = tk.StringVar()
        ttk.Label(supp_info, textvariable=self.supp_present_var).grid(row=1, column=1, sticky="w", padx=10, pady=5)

        ttk.Separator(self.supp_step3, orient="horizontal").pack(fill="x", pady=5)

        supp_draw_info = ttk.Frame(self.supp_step3)
        supp_draw_info.pack(pady=10)
        supp_draw_info.columnconfigure(0, weight=1)
        supp_draw_info.columnconfigure(1, weight=1)
        ttk.Label(supp_draw_info, text="专家姓名:").grid(row=0, column=0, sticky="e", padx=10, pady=5)
        self.supp_name_var = tk.StringVar()
        ttk.Label(supp_draw_info, textvariable=self.supp_name_var, font=("Microsoft YaHei", 14)).grid(row=0, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(supp_draw_info, text="手机号:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
        self.supp_phone_var = tk.StringVar()
        ttk.Label(supp_draw_info, textvariable=self.supp_phone_var).grid(row=1, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(supp_draw_info, text="单位:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
        self.supp_unit_var = tk.StringVar()
        ttk.Label(supp_draw_info, textvariable=self.supp_unit_var).grid(row=2, column=1, sticky="w", padx=10, pady=5)

        supp_btns = ttk.Frame(self.supp_step3)
        supp_btns.pack(pady=10)
        self.supp_btn_draw = ttk.Button(supp_btns, text="抽签", command=self._supplement_start_animation, width=12)
        self.supp_btn_stop = ttk.Button(supp_btns, text="停止", command=self._supplement_stop_animation, width=12)
        self.supp_btn_present = ttk.Button(supp_btns, text="到场", command=lambda: self._supplement_mark_result(True), width=12)
        self.supp_btn_absent = ttk.Button(supp_btns, text="不到场", command=lambda: self._supplement_mark_result(False), width=12)
        self.supp_btn_draw.grid(row=0, column=0, padx=10)
        self.supp_btn_stop.grid(row=0, column=0, padx=10)
        self.supp_btn_present.grid(row=0, column=1, padx=10)
        self.supp_btn_absent.grid(row=0, column=2, padx=10)

        supp_btn3 = ttk.Frame(self.supp_step3)
        supp_btn3.pack(pady=15)
        ttk.Button(supp_btn3, text="完成补录", command=self._supplement_finish, width=12).pack(side="right", padx=5)

    def _refresh_supplement_session_list(self):
        for i in self.supp_sessions_tree.get_children():
            self.supp_sessions_tree.delete(i)
        for s in self.app.db.get_completed_sessions():
            self.supp_sessions_tree.insert(
                "", "end",
                values=(s["id"], s["title"], s["created_at"]),
            )
        self.app.auto_adjust_columns(self.supp_sessions_tree)

    def _show_supplement_step1(self):
        self.supp_step2.pack_forget()
        self.supp_step3.pack_forget()
        self.supp_step1.pack(fill="both", expand=True)

    def _show_supplement_step2(self):
        self.supp_step1.pack_forget()
        self.supp_step3.pack_forget()
        self.supp_step2.pack(fill="both", expand=True)
        for w in self.supp_check_frame.winfo_children():
            w.destroy()
        self.supp_check_vars.clear()
        self.supp_present_logs_by_id = {}
        logs = self.app.db.get_present_logs(self.supplement_session_id)
        for r in logs:
            var = tk.BooleanVar(value=False)
            self.supp_check_vars[r["id"]] = var
            self.supp_present_logs_by_id[r["id"]] = r
            cb = ttk.Checkbutton(
                self.supp_check_frame,
                text=f"{r['name']}（{r['phone']}）",
                variable=var,
            )
            cb.pack(anchor="w", padx=5, pady=3)

    def _show_supplement_step3(self):
        self.supp_step1.pack_forget()
        self.supp_step2.pack_forget()
        self.supp_step3.pack(fill="both", expand=True)
        self.supplement_vacant_count = len([v for v in self.supp_check_vars.values() if v.get()])
        self.supplement_order_base = self.app.db.get_max_order_no(self.supp_new_session_id)
        self.supp_present_count = 0
        self.supp_current_person = None
        self.supp_order_no = self.supplement_order_base
        people = self.app.db.get_available_people()
        self.supp_people_cache = list(people)
        self.supp_drawn_person_ids = set()
        c = self.app.db.conn.cursor()
        c.execute("SELECT person_id FROM draw_logs WHERE session_id=?", (self.supp_new_session_id,))
        for row in c.fetchall():
            self.supp_drawn_person_ids.add(row["person_id"])
        self.supp_status_var.set(f"需补 {self.supplement_vacant_count} 人")
        self.supp_present_var.set(f"0 / {self.supplement_vacant_count}")
        self.supp_name_var.set("")
        self.supp_phone_var.set("")
        self.supp_unit_var.set("")
        self.supp_btn_stop.grid_remove()
        self.supp_btn_present.grid_remove()
        self.supp_btn_absent.grid_remove()
        self.supp_btn_draw.grid()
        self.supp_btn_draw["state"] = "normal"
        self.supp_btn_present["state"] = "disabled"
        self.supp_btn_absent["state"] = "disabled"

    def _supplement_on_session_selected(self):
        sel = self.supp_sessions_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个论证项目")
            return
        self.supplement_session_id = int(self.supp_sessions_tree.item(sel[0], "values")[0])
        self._show_supplement_step2()

    def _supplement_confirm_absent(self):
        """补签时标记不到场：新增记录（到场情况=后续不到场，备注=专家后续不到场，进行再次抽选。），不修改原记录"""
        selected = [log_id for log_id, var in self.supp_check_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("提示", "请至少勾选一名需改为不到场的到场人员")
            return
        c = self.app.db.conn.cursor()
        c.execute("SELECT title FROM sessions WHERE id=?", (self.supplement_session_id,))
        row = c.fetchone()
        original_title = row["title"] if row else ""
        new_title = f"（补抽）{original_title}"
        
        self.supp_new_session_id = self.app.db.create_session(new_title)
        
        selected_sorted = sorted(
            selected,
            key=lambda lid: self.supp_present_logs_by_id[lid]["order_no"] if lid in self.supp_present_logs_by_id else 0,
        )
        order_no = 1
        for log_id in selected_sorted:
            log_info = self.supp_present_logs_by_id.get(log_id)
            if log_info:
                self.app.db.add_draw_log(
                    self.supp_new_session_id,
                    log_info["person_id"],
                    order_no,
                    present=False,
                    absent_reason="专家后续不到场，进行再次抽选。",
                )
                order_no += 1
        self._show_supplement_step3()

    def _ask_reason_dialog(self, prompt="请选择不到场理由"):
        ABSENT_REASONS = [
            "工作忙，没空",
            "出差在外，无法参加",
            "身体不适，无法参加",
            "电话没人接，联系不上",
            "超出新专家上限人数",
        ]
        dlg = tk.Toplevel(self)
        dlg.title(prompt)
        dlg.grab_set()
        dlg.resizable(False, False)
        container = ttk.Frame(dlg, padding=25)
        container.pack(fill="both", expand=True)
        
        ttk.Label(container, text="请选择不到场理由：", font=SUBTITLE_FONT).pack(anchor="w", pady=(0, 15))
        
        reason_var = tk.StringVar(value=ABSENT_REASONS[0])
        for reason in ABSENT_REASONS:
            ttk.Radiobutton(
                container, 
                text=reason, 
                variable=reason_var, 
                value=reason
            ).pack(anchor="w", pady=3)
        
        res = {"value": None}

        def on_ok():
            res["value"] = reason_var.get()
            dlg.destroy()

        def on_cancel():
            res["value"] = None
            dlg.destroy()

        btn_frame = ttk.Frame(container)
        btn_frame.pack(pady=(20, 0))
        ttk.Button(btn_frame, text="确定", command=on_ok, width=11).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="取消", command=on_cancel, width=11).pack(side="left", padx=10)
        
        dlg.update_idletasks()
        width = dlg.winfo_width()
        height = dlg.winfo_height()
        x = (dlg.winfo_screenwidth() // 2) - (width // 2)
        y = (dlg.winfo_screenheight() // 2) - (height // 2)
        dlg.geometry(f"+{x}+{y}")
        
        dlg.wait_window()
        return res["value"]

    def _supplement_start_animation(self):
        if self.supp_present_count >= self.supplement_vacant_count:
            messagebox.showinfo("提示", "本次补录已完成")
            return
        
        available_people = [p for p in self.supp_people_cache if p["id"] not in self.supp_drawn_person_ids]
        if not available_people:
            messagebox.showwarning("提示", "专家名库已被抽完，无法继续抽签")
            return

        self.supp_is_drawing = True
        self.supp_btn_draw.grid_remove()
        self.supp_btn_stop.grid()
        self._supplement_animate_draw()

    def _supplement_animate_draw(self):
        if not self.supp_is_drawing:
            return
        available_people = [p for p in self.supp_people_cache if p["id"] not in self.supp_drawn_person_ids]
        if not available_people:
            self.supp_is_drawing = False
            messagebox.showwarning("提示", "专家名库已被抽完，无法继续抽签")
            self.supp_btn_stop.grid_remove()
            self.supp_btn_draw.grid()
            return
        person = random.choice(available_people)
        self.supp_name_var.set(person["name"])
        self.supp_phone_var.set(person["phone"])
        try:
            unit_value = person["unit"] if "unit" in person.keys() else ""
        except (KeyError, TypeError):
            unit_value = ""
        self.supp_unit_var.set(unit_value)
        self.supp_animation_id = self.after(50, self._supplement_animate_draw)

    def _supplement_stop_animation(self):
        if not self.supp_is_drawing:
            return
        self.supp_is_drawing = False
        if self.supp_animation_id:
            self.after_cancel(self.supp_animation_id)
            self.supp_animation_id = None

        available_people = [p for p in self.supp_people_cache if p["id"] not in self.supp_drawn_person_ids]
        if not available_people:
            messagebox.showwarning("提示", "专家名库已被抽完，无法继续抽签")
            self.supp_btn_stop.grid_remove()
            self.supp_btn_draw.grid()
            return
        
        self.supp_current_person = random.choice(available_people)
        self.supp_drawn_person_ids.add(self.supp_current_person["id"])
        self.supp_order_no += 1
        self.supp_name_var.set(self.supp_current_person["name"])
        self.supp_phone_var.set(self.supp_current_person["phone"])
        try:
            unit_value = self.supp_current_person["unit"] if "unit" in self.supp_current_person.keys() else ""
        except (KeyError, TypeError):
            unit_value = ""
        self.supp_unit_var.set(unit_value)

        self.supp_btn_stop.grid_remove()
        self.supp_btn_present.grid()
        self.supp_btn_absent.grid()
        self.supp_btn_present["state"] = "normal"
        self.supp_btn_absent["state"] = "normal"

    def _supplement_mark_result(self, present: bool):
        if self.supp_current_person is None:
            messagebox.showwarning("提示", "请先抽签")
            return
        reason = ""
        if not present:
            reason = self._ask_reason_dialog()
            if reason is None:
                return
        self.app.db.add_draw_log(
            self.supp_new_session_id,
            self.supp_current_person["id"],
            self.supp_order_no,
            present,
            reason,
        )
        if present:
            self.supp_present_count += 1
            self.supp_present_var.set(f"{self.supp_present_count} / {self.supplement_vacant_count}")
        self.supp_current_person = None
        self.supp_name_var.set("")
        self.supp_phone_var.set("")
        self.supp_unit_var.set("")
        self.supp_btn_present.grid_remove()
        self.supp_btn_absent.grid_remove()
        self.supp_btn_draw.grid()
        if self.supp_present_count >= self.supplement_vacant_count:
            self.supp_btn_draw["state"] = "disabled"
            self.supp_status_var.set("补录已完成")
            self.app.db.complete_session(self.supp_new_session_id)
            messagebox.showinfo("完成", "本次补录流程已完成")
            self.app.send_sessions_email(self, [self.supp_new_session_id])
            self._show_choice()

    def _supplement_finish(self):
        if self.supp_present_count < self.supplement_vacant_count:
            if not messagebox.askyesno("提示", "尚未补足全部名额，确定要结束吗？"):
                return
        if self.supp_new_session_id:
            self.app.db.complete_session(self.supp_new_session_id)
        self._show_choice()

    def set_buttons_state(self, started):
        self.btn_stop.grid_remove()
        self.btn_present.grid_remove()
        self.btn_absent.grid_remove()
        if not started:
            self.btn_draw["state"] = "disabled"
            self.btn_draw.grid()
            self.btn_cancel.grid()
        else:
            self.btn_draw["state"] = "normal"
            self.btn_draw.grid()
            self.btn_cancel.grid()

    def _auto_start_session(self):
        """自动开始新抽签会话"""
        people = self.app.db.get_available_people()
        if not people:
            messagebox.showwarning("提示", "当前无可用专家（可能全部被屏蔽），请先在专家名库中添加人员或取消屏蔽")
            self._show_choice()
            return

        sid = self.app.db.create_session(self.title_var.get())
        self.current_session_id = sid
        self.present_count = 0
        self.order_no = 0
        self.current_person = None
        self.people_cache = list(people)
        self.drawn_person_ids = set()
        self.present_var.set("0 / 3")
        self.name_var.set("")
        self.phone_var.set("")
        self.unit_display_var.set("")
        self.set_buttons_state(started=True)

    def _cancel_draw(self):
        """取消当前抽签，返回选择界面"""
        if messagebox.askyesno("确认", "确定要取消当前抽签吗？"):
            # 删除未完成的会话及其日志
            if self.current_session_id:
                c = self.app.db.conn.cursor()
                c.execute("DELETE FROM draw_logs WHERE session_id=?", (self.current_session_id,))
                c.execute("DELETE FROM sessions WHERE id=?", (self.current_session_id,))
                self.app.db.conn.commit()
            self._show_choice()

    def start_draw_animation(self):
        if self.current_session_id is None:
            messagebox.showwarning("提示", "抽签会话未正确初始化")
            return
        if self.present_count >= 3:
            messagebox.showinfo("提示", "本次抽签已完成 3 名到场人员")
            return
        
        available_people = [p for p in self.people_cache if p["id"] not in self.drawn_person_ids]
        if not available_people:
            messagebox.showwarning("提示", "专家名库已被抽完，无法继续抽签")
            return

        self.is_drawing = True
        self.btn_draw.grid_remove()
        self.btn_stop.grid()
        self._animate_draw()

    def _animate_draw(self):
        if not self.is_drawing:
            return
        available_people = [p for p in self.people_cache if p["id"] not in self.drawn_person_ids]
        if not available_people:
            self.is_drawing = False
            messagebox.showwarning("提示", "专家名库已被抽完，无法继续抽签")
            self.btn_stop.grid_remove()
            self.btn_draw.grid()
            return
        person = random.choice(available_people)
        self.name_var.set(person["name"])
        self.phone_var.set(person["phone"])
        try:
            unit_value = person["unit"] if "unit" in person.keys() else ""
        except (KeyError, TypeError):
            unit_value = ""
        self.unit_display_var.set(unit_value)
        self.draw_animation_id = self.after(50, self._animate_draw)

    def stop_draw_animation(self):
        if not self.is_drawing:
            return
        self.is_drawing = False
        if self.draw_animation_id:
            self.after_cancel(self.draw_animation_id)
            self.draw_animation_id = None

        available_people = [p for p in self.people_cache if p["id"] not in self.drawn_person_ids]
        if not available_people:
            messagebox.showwarning("提示", "专家名库已被抽完，无法继续抽签")
            self.btn_stop.grid_remove()
            self.btn_draw.grid()
            return
        
        self.current_person = random.choice(available_people)
        self.drawn_person_ids.add(self.current_person["id"])
        self.order_no += 1
        self.name_var.set(self.current_person["name"])
        self.phone_var.set(self.current_person["phone"])
        try:
            unit_value = self.current_person["unit"] if "unit" in self.current_person.keys() else ""
        except (KeyError, TypeError):
            unit_value = ""
        self.unit_display_var.set(unit_value)

        self.btn_stop.grid_remove()
        self.btn_present.grid()
        self.btn_absent.grid()

    def mark_result(self, present: bool):
        if self.current_person is None:
            messagebox.showwarning("提示", "请先抽签")
            return
        reason = ""
        if not present:
            reason = self._ask_reason()
            if reason is None:
                return

        self.app.db.add_draw_log(
            self.current_session_id,
            self.current_person["id"],
            self.order_no,
            present,
            reason,
        )

        if present:
            self.present_count += 1
            self.present_var.set(f"{self.present_count} / 3")

        self.current_person = None
        self.name_var.set("")
        self.phone_var.set("")
        self.unit_display_var.set("")
        self.btn_present.grid_remove()
        self.btn_absent.grid_remove()
        self.btn_draw.grid()

        if self.present_count >= 3:
            self.btn_draw["state"] = "disabled"
            # 标记会话为已完成
            self.app.db.complete_session(self.current_session_id)
            messagebox.showinfo("完成", "本次抽签流程已完成 3 名到场人员")
            self.app.send_sessions_email(self, [self.current_session_id])
            # 抽签完成后自动返回抽签功能界面
            self._show_choice()

    def _ask_reason(self):
        ABSENT_REASONS = [
            "工作忙，没空",
            "出差在外，无法参加",
            "身体不适，无法参加",
            "电话没人接，联系不上",
            "超出新专家上限人数",
        ]
        dlg = tk.Toplevel(self)
        dlg.title("请选择不到场理由")
        dlg.grab_set()
        dlg.resizable(False, False)

        container = ttk.Frame(dlg, padding=25)
        container.pack(fill="both", expand=True)
        
        ttk.Label(container, text="请选择不到场理由：", font=SUBTITLE_FONT).pack(anchor="w", pady=(0, 15))
        
        reason_var = tk.StringVar(value=ABSENT_REASONS[0])
        for reason in ABSENT_REASONS:
            ttk.Radiobutton(
                container, 
                text=reason, 
                variable=reason_var, 
                value=reason
            ).pack(anchor="w", pady=3)

        res = {"value": None}

        def on_ok():
            res["value"] = reason_var.get()
            dlg.destroy()

        def on_cancel():
            res["value"] = None
            dlg.destroy()

        btn_frame = ttk.Frame(container)
        btn_frame.pack(pady=(20, 0))
        ttk.Button(btn_frame, text="确定", command=on_ok, width=11).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="取消", command=on_cancel, width=11).pack(side="left", padx=10)
        
        dlg.update_idletasks()
        width = dlg.winfo_width()
        height = dlg.winfo_height()
        x = (dlg.winfo_screenwidth() // 2) - (width // 2)
        y = (dlg.winfo_screenheight() // 2) - (height // 2)
        dlg.geometry(f"+{x}+{y}")

        dlg.wait_window()
        return res["value"]


class UsersFrame(ttk.Frame):
    """账户管理：仅管理员使用"""

    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.build_ui()
        self.refresh()

    def build_ui(self):
        outer = ttk.Frame(self, padding=30)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x", pady=(0, 15))

        ttk.Label(top, text="账户管理（仅管理员）", font=SUBTITLE_FONT).pack(
            side="left", padx=8
        )

        # 标题下方中部按钮栏
        btn_bar = ttk.Frame(outer)
        btn_bar.pack(pady=15)

        self.btn_add = ttk.Button(btn_bar, text="新增账户", command=self.add_user, width=10)
        self.btn_edit = ttk.Button(btn_bar, text="编辑账户", command=self.edit_user, width=10)
        self.btn_del = ttk.Button(btn_bar, text="删除账户", command=self.delete_user, width=10)
        self.btn_set_admin = ttk.Button(
            btn_bar, text="设为管理员", command=lambda: self.change_role("admin"), width=11
        )
        self.btn_set_user = ttk.Button(
            btn_bar, text="设为普通用户", command=lambda: self.change_role("user"), width=12
        )

        for b in (
                self.btn_add,
                self.btn_edit,
                self.btn_del,
                self.btn_set_admin,
                self.btn_set_user,
        ):
            b.pack(side="left", padx=6)

        table_frame = ttk.Frame(outer)
        table_frame.pack(fill="both", expand=True)

        columns = ("id", "username", "email", "role")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=18,
        )
        # 隐藏 ID 列，仅用于内部
        self.tree.heading("id", text="")
        self.tree.heading("username", text="用户名")
        self.tree.heading("email", text="电子邮箱")
        self.tree.heading("role", text="角色")

        self.tree.column("id", width=0, minwidth=0, anchor="center", stretch=False)
        self.tree.column("username", width=180, anchor="w", stretch=True)
        self.tree.column("email", width=250, anchor="w", stretch=True)
        self.tree.column("role", width=120, anchor="center", stretch=False)
        self.tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")



    def refresh(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for row in self.app.db.get_all_users():
            role_text = "管理员" if row["role"] == "admin" else "普通用户"
            self.tree.insert(
                "",
                "end",
                values=(row["id"], row["username"], row["email"] or "", role_text),
            )
        self.app.auto_adjust_columns(self.tree)

    def _get_selected_id_role(self):
        sel = self.tree.selection()
        if not sel:
            return None, None
        values = self.tree.item(sel[0], "values")
        uid = int(values[0])
        role_text = values[3]
        role = "admin" if role_text == "管理员" else "user"
        return uid, role

    def add_user(self):
        self._edit_dialog()

    def edit_user(self):
        uid, role = self._get_selected_id_role()
        if not uid:
            messagebox.showwarning("提示", "请先选择一个账户")
            return
        # 从 db 获取该用户
        users = {u["id"]: u for u in self.app.db.get_all_users()}
        if uid not in users:
            return
        u = users[uid]
        self._edit_dialog(uid, u["username"], u["role"], u["email"] or "")

    def _edit_dialog(self, user_id=None, username="", role="user", email=""):
        dlg = tk.Toplevel(self)
        dlg.title("账户信息")
        dlg.grab_set()
        dlg.resizable(False, False)

        container = ttk.Frame(dlg, padding=25)
        container.grid(row=0, column=0, sticky="nsew")
        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)

        ttk.Label(container, text="用户名:").grid(
            row=0, column=0, padx=8, pady=10, sticky="e"
        )
        username_var = tk.StringVar(value=username)
        ttk.Entry(container, textvariable=username_var, width=28).grid(
            row=0, column=1, padx=8, pady=10, sticky="w"
        )

        ttk.Label(container, text="密码:").grid(
            row=1, column=0, padx=8, pady=10, sticky="e"
        )
        pwd_var = tk.StringVar()
        ttk.Entry(container, textvariable=pwd_var, show="*", width=28).grid(
            row=1, column=1, padx=8, pady=10, sticky="w"
        )
        ttk.Label(
            container, text="（留空则不修改密码）", foreground="gray"
        ).grid(row=2, column=1, padx=8, pady=(0, 10), sticky="w")

        ttk.Label(container, text="电子邮箱:").grid(
            row=3, column=0, padx=8, pady=10, sticky="e"
        )
        email_var = tk.StringVar(value=email)
        ttk.Entry(container, textvariable=email_var, width=28).grid(
            row=3, column=1, padx=8, pady=10, sticky="w"
        )

        ttk.Label(container, text="角色:").grid(
            row=4, column=0, padx=8, pady=10, sticky="e"
        )
        role_var = tk.StringVar(value="管理员" if role == "admin" else "普通用户")
        cb = ttk.Combobox(
            container,
            textvariable=role_var,
            values=["管理员", "普通用户"],
            state="readonly",
            width=25,
        )
        cb.grid(row=4, column=1, padx=8, pady=10, sticky="w")

        def on_ok():
            name = username_var.get().strip()
            pwd = pwd_var.get().strip()
            email_val = email_var.get().strip()
            role_choice = "admin" if role_var.get() == "管理员" else "user"
            if not name:
                messagebox.showwarning("提示", "用户名不能为空")
                return
            try:
                if user_id is None:
                    # 新增账户时密码必填
                    if not pwd:
                        messagebox.showwarning("提示", "新增账户时密码不能为空")
                        return
                    ok, msg = self.app.db.register_user(name, pwd, role=role_choice, email=email_val)
                    if not ok:
                        messagebox.showerror("失败", msg)
                        return
                else:
                    if pwd:
                        self.app.db.update_user(user_id, name, role_choice, pwd, email=email_val)
                    else:
                        self.app.db.update_user(user_id, name, role_choice, email=email_val)
                self.refresh()
                dlg.destroy()
            except Exception as e:
                messagebox.showerror("错误", str(e))

        btn_frame = ttk.Frame(container)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(18, 0))
        ttk.Button(btn_frame, text="确定", command=on_ok, width=11).grid(
            row=0, column=0, padx=10
        )
        ttk.Button(btn_frame, text="取消", command=dlg.destroy, width=11).grid(
            row=0, column=1, padx=10
        )

    def delete_user(self):
        uid, role = self._get_selected_id_role()
        if not uid:
            messagebox.showwarning("提示", "请先选择一个账户")
            return
        if not messagebox.askyesno("确认", "确定要删除该账户吗？"):
            return
        try:
            self.app.db.delete_user(uid, current_user_id=self.app.current_user["id"])
            self.refresh()
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def change_role(self, new_role):
        uid, _old_role = self._get_selected_id_role()
        if not uid:
            messagebox.showwarning("提示", "请先选择一个账户")
            return
        users = {u["id"]: u for u in self.app.db.get_all_users()}
        if uid not in users:
            return
        u = users[uid]
        try:
            self.app.db.update_user(uid, u["username"], new_role)
            self.refresh()
        except Exception as e:
            messagebox.showerror("错误", str(e))


class MailConfigFrame(ttk.Frame):
    """设置（仅管理员）"""

    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.build_ui()
        self.load_config()

    def build_ui(self):
        outer = ttk.Frame(self, padding=30)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="设置（仅管理员）", font=SUBTITLE_FONT).pack(
            anchor="w", padx=5, pady=(0, 15)
        )
        
        # 邮件设置区域
        mail_frame = ttk.LabelFrame(outer, text="邮件设置", padding=15)
        mail_frame.pack(fill="x", pady=(0, 10), padx=10)

        form = ttk.Frame(mail_frame)
        form.pack(pady=10, padx=10, anchor="center")

        for i in range(2):
            form.columnconfigure(i, weight=1)

        ttk.Label(form, text="SMTP 服务器:").grid(row=0, column=0, sticky="e", padx=8, pady=6)
        self.smtp_server_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.smtp_server_var, width=32).grid(
            row=0, column=1, sticky="w", padx=8, pady=6
        )

        ttk.Label(form, text="端口:").grid(row=1, column=0, sticky="e", padx=8, pady=6)
        self.smtp_port_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.smtp_port_var, width=10).grid(
            row=1, column=1, sticky="w", padx=8, pady=6
        )

        ttk.Label(form, text="发件人邮箱:").grid(row=2, column=0, sticky="e", padx=8, pady=6)
        self.from_addr_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.from_addr_var, width=32).grid(
            row=2, column=1, sticky="w", padx=8, pady=6
        )

        ttk.Label(form, text="登录用户名:").grid(row=3, column=0, sticky="e", padx=8, pady=6)
        self.username_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.username_var, width=32).grid(
            row=3, column=1, sticky="w", padx=8, pady=6
        )

        ttk.Label(form, text="登录密码:").grid(row=4, column=0, sticky="e", padx=8, pady=6)
        self.password_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.password_var, show="*", width=32).grid(
            row=4, column=1, sticky="w", padx=8, pady=6
        )

        self.use_ssl_var = tk.BooleanVar(value=True)
        self.use_tls_var = tk.BooleanVar(value=False)

        ssl_frame = ttk.Frame(mail_frame)
        ssl_frame.pack(pady=5)
        ttk.Checkbutton(
            ssl_frame, text="使用 SSL", variable=self.use_ssl_var
        ).pack(side="left", padx=10)
        ttk.Checkbutton(
            ssl_frame, text="使用 STARTTLS", variable=self.use_tls_var
        ).pack(side="left", padx=10)

        btn_frame = ttk.Frame(mail_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="保存配置", command=self.save_config, width=12).pack(
            side="left", padx=10
        )
        
        # 接收邮箱列表区域（放在邮件设置内）
        ttk.Separator(mail_frame, orient="horizontal").pack(fill="x", pady=10)
        
        email_list_label = ttk.Label(mail_frame, text="接收邮箱列表", font=("Microsoft YaHei", 10, "bold"))
        email_list_label.pack(anchor="w", padx=5, pady=(0, 5))
        
        # 按钮栏（居中）
        email_btn_frame = ttk.Frame(mail_frame)
        email_btn_frame.pack(pady=5)
        ttk.Button(email_btn_frame, text="添加邮箱", command=self._add_email, width=10).pack(side="left", padx=5)
        ttk.Button(email_btn_frame, text="编辑邮箱", command=self._edit_email, width=10).pack(side="left", padx=5)
        ttk.Button(email_btn_frame, text="删除邮箱", command=self._delete_email, width=10).pack(side="left", padx=5)
        
        # 邮箱列表
        list_frame = ttk.Frame(mail_frame)
        list_frame.pack(fill="both", expand=True, pady=5)
        
        columns = ("id", "email", "name")
        self.email_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=6)
        self.email_tree.heading("id", text="")
        self.email_tree.heading("email", text="邮箱地址")
        self.email_tree.heading("name", text="名称")
        self.email_tree.column("id", width=0, minwidth=0, stretch=False)
        self.email_tree.column("email", width=250, anchor="w")
        self.email_tree.column("name", width=150, anchor="w")
        self.email_tree.pack(side="left", fill="both", expand=True)
        
        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.email_tree.yview)
        self.email_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        
        self._refresh_email_list()
        
        # 重置设置区域
        reset_frame = ttk.LabelFrame(outer, text="重置设置", padding=15)
        reset_frame.pack(fill="x", pady=10, padx=10)
        
        # 清空专家名库
        row1 = ttk.Frame(reset_frame)
        row1.pack(fill="x", pady=5)
        ttk.Label(row1, text="清空当前专家名库：").pack(side="left", padx=10)
        ttk.Button(row1, text="清空名库", command=self._clear_people, width=12).pack(side="right", padx=10)

    def _refresh_email_list(self):
        for i in self.email_tree.get_children():
            self.email_tree.delete(i)
        for r in self.app.db.get_email_recipients():
            self.email_tree.insert("", "end", values=(r["id"], r["email"], r["name"] or ""))
        self.app.auto_adjust_columns(self.email_tree)

    def _add_email(self):
        self._edit_email_dialog()

    def _edit_email(self):
        sel = self.email_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个邮箱")
            return
        values = self.email_tree.item(sel[0], "values")
        self._edit_email_dialog(int(values[0]), values[1], values[2])

    def _edit_email_dialog(self, recipient_id=None, email="", name=""):
        dlg = tk.Toplevel(self)
        dlg.title("添加邮箱" if recipient_id is None else "编辑邮箱")
        dlg.grab_set()
        dlg.resizable(False, False)

        container = ttk.Frame(dlg, padding=25)
        container.grid(row=0, column=0, sticky="nsew")
        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)

        ttk.Label(container, text="邮箱地址:").grid(
            row=0, column=0, padx=8, pady=10, sticky="e"
        )
        email_var = tk.StringVar(value=email)
        ttk.Entry(container, textvariable=email_var, width=28).grid(
            row=0, column=1, padx=8, pady=10, sticky="w"
        )

        ttk.Label(container, text="名称:").grid(
            row=1, column=0, padx=8, pady=10, sticky="e"
        )
        name_var = tk.StringVar(value=name)
        ttk.Entry(container, textvariable=name_var, width=28).grid(
            row=1, column=1, padx=8, pady=10, sticky="w"
        )

        def on_ok():
            email_val = email_var.get().strip()
            name_val = name_var.get().strip()
            if not email_val:
                messagebox.showwarning("提示", "邮箱地址不能为空")
                return
            try:
                if recipient_id is None:
                    self.app.db.add_email_recipient(email_val, name_val)
                else:
                    self.app.db.update_email_recipient(recipient_id, email_val, name_val)
                self._refresh_email_list()
                dlg.destroy()
            except Exception as e:
                messagebox.showerror("错误", str(e))

        btn_frame = ttk.Frame(container)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=(18, 0))
        ttk.Button(btn_frame, text="确定", command=on_ok, width=11).grid(
            row=0, column=0, padx=10
        )
        ttk.Button(btn_frame, text="取消", command=dlg.destroy, width=11).grid(
            row=0, column=1, padx=10
        )

        dlg.update_idletasks()
        width = dlg.winfo_width()
        height = dlg.winfo_height()
        x = (dlg.winfo_screenwidth() // 2) - (width // 2)
        y = (dlg.winfo_screenheight() // 2) - (height // 2)
        dlg.geometry(f"+{x}+{y}")

    def _delete_email(self):
        sel = self.email_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个邮箱")
            return
        values = self.email_tree.item(sel[0], "values")
        if messagebox.askyesno("确认", f"确定要删除邮箱 {values[1]} 吗？"):
            self.app.db.delete_email_recipient(int(values[0]))
            self._refresh_email_list()

    def _clear_people(self):
        if messagebox.askyesno("确认", "确定要清空专家名库吗？此操作不可恢复。"):
            self.app.db.delete_all_people()
            messagebox.showinfo("成功", "专家名库已清空")

    def load_config(self):
        cfg = self.app.db.get_mail_config()
        self.smtp_server_var.set(cfg.get("smtp_server") or "")
        self.smtp_port_var.set(str(cfg.get("smtp_port") or ""))
        self.from_addr_var.set(cfg.get("from_addr") or "")
        self.username_var.set(cfg.get("username") or "")
        self.password_var.set(cfg.get("password") or "")
        self.use_ssl_var.set(cfg.get("use_ssl"))
        self.use_tls_var.set(cfg.get("use_tls"))

    def save_config(self):
        try:
            port = int(self.smtp_port_var.get() or 0)
        except ValueError:
            messagebox.showerror("错误", "端口必须是数字")
            return
        cfg = {
            "smtp_server": self.smtp_server_var.get().strip(),
            "smtp_port": port,
            "from_addr": self.from_addr_var.get().strip(),
            "username": self.username_var.get().strip(),
            "password": self.password_var.get(),
            "use_ssl": self.use_ssl_var.get(),
            "use_tls": self.use_tls_var.get(),
        }
        self.app.db.save_mail_config(cfg)
        messagebox.showinfo("成功", "邮件配置已保存")


class MainFrame(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.account_tab_added = False
        self.mail_tab_added = False
        self.build_ui()

    def build_ui(self):
        outer = ttk.Frame(self, padding=30)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x", pady=(0, 15))

        self.user_label = ttk.Label(top, text="")
        self.user_label.pack(side="left", padx=8)
        
        # 退出登录按钮（放在当前用户右边）
        self.btn_logout = ttk.Button(top, text="退出登录", command=self.logout, width=10)
        self.btn_logout.pack(side="left", padx=5)
        
        # 右上角小logo
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "珠海安防协会logo.gif")
            if os.path.exists(logo_path):
                logo_img = tk.PhotoImage(file=logo_path)
                logo_img = logo_img.subsample(16, 16)  # 缩小到1/16
                logo_label = ttk.Label(top, image=logo_img)
                logo_label.image = logo_img
                logo_label.pack(side="right", padx=10)
        except Exception:
            pass

        self.notebook = ttk.Notebook(outer)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=5)

        self.people_frame = PeopleFrame(self.notebook, self.app)
        self.draw_frame = DrawFrame(self.notebook, self.app)
        self.logs_frame = LogsFrame(self.notebook, self.app)
        self.account_frame = UsersFrame(self.notebook, self.app)
        self.mail_frame = MailConfigFrame(self.notebook, self.app)

        self.notebook.add(self.draw_frame, text="抽签")
        self.notebook.add(self.people_frame, text="专家名库")
        self.notebook.add(self.logs_frame, text="日志")
        # 账户管理 / 邮件设置 Tab 在管理员登录后动态添加
        
        # 绑定 tab 切换事件，切换到日志界面时自动刷新
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def refresh_user_info(self):
        u = self.app.current_user
        if u:
            text = f"当前用户：{u['username']}（{'管理员' if u['role']=='admin' else '普通用户'}）"
            self.user_label.configure(text=text)
            self.people_frame.set_admin_mode(u["role"] == "admin")

            # 管理员才显示账户管理 & 邮件设置 Tab
            if u["role"] == "admin":
                if not self.account_tab_added:
                    self.notebook.add(self.account_frame, text="账户管理")
                    self.account_tab_added = True
                if not self.mail_tab_added:
                    self.notebook.add(self.mail_frame, text="设置")
                    self.mail_tab_added = True
            else:
                if self.account_tab_added:
                    self.notebook.forget(self.account_frame)
                    self.account_tab_added = False
                if self.mail_tab_added:
                    self.notebook.forget(self.mail_frame)
                    self.mail_tab_added = False
        else:
            self.user_label.configure(text="")
            if self.account_tab_added:
                self.notebook.forget(self.account_frame)
                self.account_tab_added = False
            if self.mail_tab_added:
                self.notebook.forget(self.mail_frame)
                self.mail_tab_added = False

    def _on_tab_changed(self, event):
        """Tab 切换事件处理，切换到日志界面时自动刷新"""
        current_tab = self.notebook.index(self.notebook.select())
        # 日志界面是第3个tab（索引为2）
        if current_tab == 2:
            self.logs_frame.refresh_sessions()

    def logout(self):
        if messagebox.askyesno("确认", "确定要退出登录吗？"):
            self.app.current_user = None
            self.app.show_login_frame()


# -------------------- 应用主类 --------------------


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        # 设置窗口图标
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "珠海安防协会logo.gif")
            if os.path.exists(logo_path):
                icon_img = tk.PhotoImage(file=logo_path)
                icon_img = icon_img.subsample(16, 16)  # 缩小图标
                self.iconphoto(False, icon_img)
        except Exception:
            pass

        # 初始窗口大小，允许缩放
        self.title("珠海安防协会项目论证专家抽签系统")
        self.geometry("1400x900")
        self.minsize(1600, 1150)
        self.resizable(True, True)

        # 全局字体
        self.option_add("*Font", BASE_FONT)

        # Windows 下使用较为现代的 ttk 主题
        self.style = ttk.Style()
        try:
            if os.name == "nt":
                self.style.theme_use("vista")
        except Exception:
            pass

        # 放大 Notebook Tab 样式（选项卡标签）
        self.style.configure(
            "TNotebook.Tab",
            font=("Microsoft YaHei", 12, "bold"),
            padding=(20, 8),
        )

        # 全局按钮默认内边距稍大
        self.style.configure("TButton", padding=(12, 6))

        # Treeview 行高与字体（根据字体实际高度动态计算）
        font_obj = tkfont.Font(family=BASE_FONT[0], size=BASE_FONT[1])
        font_metrics = font_obj.metrics()
        row_height = font_metrics["linespace"] + 10
        self.style.configure(
            "Treeview",
            rowheight=row_height,
            font=BASE_FONT,
        )
        self.style.configure(
            "Treeview.Heading",
            font=("Microsoft YaHei", 11, "bold"),
        )

        self.db = Database()
        self.current_user = None

        self.container = ttk.Frame(self, padding=20)
        self.container.pack(fill="both", expand=True)

        self.login_frame = LoginFrame(self.container, self)
        self.register_frame = RegisterFrame(self.container, self)
        self.main_frame = MainFrame(self.container, self)

        self.show_login_frame()

        # 在 Tk 初始化完成后，根据实际 DPI 设置 tk 的 scaling
        self.after(0, self._adjust_scaling_for_dpi)

    def send_sessions_email(self, parent, session_ids):
        """将一个或多个论证项目的日志发送到所有管理员邮箱"""
        cfg = self.db.get_mail_config()
        if not cfg["smtp_server"] or not cfg["from_addr"]:
            messagebox.showerror("错误", "请先在「邮件设置」中配置 SMTP 服务器和发件人邮箱", parent=parent)
            return False
        recipient_emails = self.db.get_recipient_emails()
        if not recipient_emails:
            messagebox.showerror("错误", "当前没有配置接收邮箱，请在「设置」中添加接收邮箱", parent=parent)
            return False

        # 组装邮件内容
        lines = []
        for sid in session_ids:
            sessions = [s for s in self.db.get_sessions() if s["id"] == sid]
            if not sessions:
                continue
            s = sessions[0]
            lines.append(f"论证项目 ID: {s['id']}")
            lines.append(f"论证项目名称: {s['title'] or ''}")
            lines.append(f"创建时间: {s['created_at']}")
            lines.append("-" * 40)
            logs = self.db.get_session_logs(sid)
            for r in logs:
                if r["present"] == 1:
                    state = "到场"
                elif r["absent_reason"] == "专家后续不到场，进行再次抽选。":
                    state = "后续不到场"
                else:
                    state = "不到场"
                lines.append(
                    f"[{r['created_at']}] 姓名:{r['name']} 手机:{r['phone']} 到场情况:{state} 备注:{r['absent_reason'] or ''}"
                )
            lines.append("=" * 60)
            lines.append("")

        if not lines:
            messagebox.showwarning("提示", "未找到可发送的日志记录", parent=parent)
            return False

        body = "\n".join(lines)
        subject = "论证项目抽签结果"

        while True:
            try:

                if cfg["use_ssl"]:
                    server = smtplib.SMTP_SSL(cfg["smtp_server"], cfg["smtp_port"] or 465, timeout=10)
                else:
                    server = smtplib.SMTP(cfg["smtp_server"], cfg["smtp_port"] or 25, timeout=10)

                try:
                    if cfg["use_tls"]:
                        server.starttls()
                    if cfg["username"]:
                        server.login(cfg["username"], cfg["password"] or "")

                    msg = EmailMessage()
                    msg["Subject"] = subject
                    msg["From"] = cfg["from_addr"]
                    msg["To"] = ", ".join(recipient_emails)
                    msg.set_content(body)

                    server.send_message(msg)
                finally:
                    server.quit()

                messagebox.showinfo("发送成功", "抽签结果邮件已成功发送至管理员邮箱！")
                return True
            except smtplib.SMTPAuthenticationError as e:
                error_msg = f"邮件发送失败：邮箱认证错误\n\n原因：用户名或密码错误\n详细信息：{e}\n\n请检查邮件设置中的用户名和密码是否正确。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except smtplib.SMTPConnectError as e:
                error_msg = f"邮件发送失败：无法连接到邮件服务器\n\n原因：服务器连接失败\n详细信息：{e}\n\n请检查SMTP服务器地址和端口是否正确。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except smtplib.SMTPServerDisconnected as e:
                error_msg = f"邮件发送失败：服务器连接断开\n\n原因：连接被服务器意外关闭\n详细信息：{e}\n\n请检查网络连接是否稳定。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except smtplib.SMTPRecipientsRefused as e:
                error_msg = f"邮件发送失败：收件人地址被拒绝\n\n原因：管理员邮箱地址无效或被服务器拒绝\n详细信息：{e}\n\n请检查管理员邮箱配置。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except smtplib.SMTPSenderRefused as e:
                error_msg = f"邮件发送失败：发件人地址被拒绝\n\n原因：发件人邮箱地址无效或未被授权\n详细信息：{e}\n\n请检查邮件设置中的发件人邮箱。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except smtplib.SMTPDataError as e:
                error_msg = f"邮件发送失败：邮件数据错误\n\n原因：邮件内容格式有问题\n详细信息：{e}\n\n请联系技术支持。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except socket.timeout as e:
                error_msg = f"邮件发送失败：连接超时\n\n原因：服务器响应时间过长\n详细信息：{e}\n\n请检查网络连接或稍后重试。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except socket.gaierror as e:
                error_msg = f"邮件发送失败：网络地址解析失败\n\n原因：无法解析SMTP服务器地址\n详细信息：{e}\n\n请检查服务器地址是否正确，以及网络连接是否正常。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except ConnectionRefusedError as e:
                error_msg = f"邮件发送失败：连接被拒绝\n\n原因：服务器拒绝连接\n详细信息：{e}\n\n请检查SMTP端口是否正确。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False
            except Exception as e:
                error_type = type(e).__name__
                error_msg = f"邮件发送失败：未知错误\n\n错误类型：{error_type}\n详细信息：{e}\n\n请检查网络和邮件配置后重试。"
                retry = messagebox.askretrycancel("发送失败", error_msg)
                if not retry:
                    return False

    def _adjust_scaling_for_dpi(self):
        try:
            pixels_per_inch = self.winfo_fpixels("1i")
            scaling = pixels_per_inch / 72.0
            scaling = max(1.0, min(scaling, 2.0))
            self.tk.call("tk", "scaling", scaling)
        except Exception:
            pass

    @staticmethod
    def auto_adjust_columns(tree, padding=20, last_column_stretch=True):
        """
        自适应调整Treeview列宽
        :param tree: Treeview控件
        :param padding: 每列内边距
        :param last_column_stretch: 是否将剩余宽度分配给最后一列
        """
        tree.update_idletasks()
        
        columns = tree["columns"]
        if not columns:
            return
        
        font_obj = tkfont.Font(family=BASE_FONT[0], size=BASE_FONT[1])
        col_widths = {}
        visible_columns = []
        
        for col in columns:
            current_width = tree.column(col, "width")
            if current_width == 0:
                col_widths[col] = 0
                continue
            visible_columns.append(col)
            heading_text = tree.heading(col, option="text") or ""
            heading_width = font_obj.measure(heading_text) + padding
            col_widths[col] = heading_width
        
        for item in tree.get_children():
            values = tree.item(item, "values")
            for i, col in enumerate(columns):
                if col not in visible_columns:
                    continue
                if i < len(values):
                    cell_text = str(values[i]) if values[i] is not None else ""
                    cell_width = font_obj.measure(cell_text) + padding
                    if cell_width > col_widths[col]:
                        col_widths[col] = cell_width
        
        total_width = sum(col_widths.values())
        
        tree_width = tree.winfo_width()
        if tree_width <= 1:
            tree_width = tree.winfo_reqwidth()
        
        if last_column_stretch and visible_columns and tree_width > total_width + 20:
            last_col = visible_columns[-1]
            extra_width = tree_width - total_width - 20
            col_widths[last_col] += extra_width
        
        for col in columns:
            if col_widths[col] == 0:
                tree.column(col, width=0, minwidth=0)
            else:
                tree.column(col, width=col_widths[col], minwidth=50)

    def _show_frame(self, frame):
        for w in self.container.winfo_children():
            w.pack_forget()
        frame.pack(fill="both", expand=True)

    def show_login_frame(self):
        self._show_frame(self.login_frame)

    def show_register_frame(self):
        self._show_frame(self.register_frame)

    def show_main_frame(self):
        self.main_frame.refresh_user_info()
        self._show_frame(self.main_frame)


# -------------------- 入口 --------------------

if __name__ == "__main__":
    set_dpi_awareness()
    app = App()
    app.mainloop()